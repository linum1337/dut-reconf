import openpyxl
path = "Команды.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_read = sheet_obj.cell(row=1, column=1)
new_pwd = sheet_obj.cell(row=1, column=2)
print(cell_read.value)
new_pwd.value = "hhh"

print(sheet_obj.cell(row=1, column=2).value)
